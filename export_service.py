"""
export_service.py
==================
Generación de reportes exportables en Excel (.xlsx) — contactos,
campañas y overview del Panel Admin. Devuelve bytes en memoria (BytesIO)
para servir directo como descarga, sin tocar disco.
"""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _style_header(ws, ncols: int):
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = "A2"


def _autosize(ws, ncols: int):
    for col in range(1, ncols + 1):
        letter = ws.cell(row=1, column=col).column_letter
        maxlen = max(
            [len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)]
            or [10]
        )
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 50)


def contacts_to_xlsx(contacts: list, field_defs: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    headers = ["Nombre", "Teléfono", "Etiqueta", "Dado de baja"] + [f.label for f in field_defs]
    ws.append(headers)
    for c in contacts:
        fields = c.get_fields()
        row = [c.name, c.phone, c.tag or "", "Sí" if c.opted_out else "No"]
        row += [fields.get(f.key, "") for f in field_defs]
        ws.append(row)
    _style_header(ws, len(headers))
    _autosize(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def campaign_report_to_xlsx(campaign, messages: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de campaña"
    ws.append([f"Campaña: {campaign.name}"])
    ws.append([f"Generado: {datetime.utcnow().isoformat(timespec='seconds')} UTC"])
    ws.append([])
    headers = ["Contacto", "Teléfono", "Estado", "Enviado el", "Error"]
    ws.append(headers)
    for m in messages:
        ws.append([
            m.contact_name, m.contact_phone, m.status,
            m.sent_at.isoformat(timespec="seconds") if m.sent_at else "",
            m.error_msg or "",
        ])
    _style_header_at(ws, 4, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _style_header_at(ws, row_idx: int, ncols: int):
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font


def clients_overview_to_xlsx(clients: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    headers = ["Empresa", "Email", "Rubro", "WhatsApp", "Contactos", "Campañas", "Agentes", "Alerta reciente"]
    ws.append(headers)
    for c in clients:
        ws.append([
            c["name"], c["email"], c.get("industry") or "",
            "Conectado" if c["whatsapp_connected"] else "Sin conectar",
            c["contacts"], c["campaigns"], c["agents"],
            c.get("last_whatsapp_error") or "",
        ])
    _style_header(ws, len(headers))
    _autosize(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
